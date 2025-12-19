from PyQt6.QtWidgets import (
    QMessageBox, QFileDialog, QDialog, QVBoxLayout, QHBoxLayout, 
    QComboBox, QLabel, QPushButton, QTableWidget, QTableWidgetItem, QHeaderView
)
from PyQt6.QtCore import QDate, Qt
from PyQt6.QtGui import QFont
from views.reports import GenerateReportDialog, ViewReportDialog
from models.purchase import PurchaseModel, ItemModel, DamageModel
from datetime import datetime, timedelta
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch


class ReportsController:
    def __init__(self, view, model):
        self.view = view
        self.model = model
        
        # Connect action buttons - updated button names
        self.view.btn_view_reports.clicked.connect(self.handle_stock_levels)
        self.view.btn_generate_reports.clicked.connect(self.handle_view_reports)
        self.view.btn_summary.clicked.connect(self.handle_low_stock_alert)
        self.view.btn_export.clicked.connect(self.handle_export_report)
        
        # Connect filters
        self.view.year_filter.currentTextChanged.connect(self.handle_filter_change)
        self.view.month_filter.currentTextChanged.connect(self.handle_filter_change)
    
    def set_user_info(self, user_name, user_role, department):
        """Set current user information."""
        self.view.current_user = user_name
        self.view.current_role = user_role
        self.view.current_department = department
        
        # Load initial charts with department filter
        self.refresh_charts()
    
    def handle_filter_change(self):
        """Handle year/month filter changes."""
        self.refresh_charts()
    
    def refresh_charts(self):
        """Refresh chart data based on current filters and department."""
        try:
            # Fetch all inventory data for bar chart
            inventory_data = self.get_all_stock_levels()
            
            # Update the chart in the view
            self.view.update_charts(inventory_data)
            
        except Exception as e:
            print(f"Error refreshing charts: {e}")
            import traceback
            traceback.print_exc()
    
    def get_all_stock_levels(self):
        """Get all stock levels for the bar chart."""
        try:
            item_model = ItemModel()
            items = item_model.list_items()
            
            # Filter by department if Department role
            if self.view.current_role == "Department" and self.view.current_department:
                items = [item for item in items 
                        if item.get('category', '').lower() == self.view.current_department.lower()]
            
            # Return all items with their stock data
            return items
            
        except Exception as e:
            print(f"Error getting all stock levels: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def handle_stock_levels(self):
        """Show stock levels report."""
        msg = QMessageBox(self.view)
        msg.setIcon(QMessageBox.Icon.Information)
        msg.setWindowTitle("Stock Levels")
        msg.setText("Stock levels are displayed in the charts below.")
        msg.setStyleSheet("""
            QMessageBox { background-color: white; }
            QLabel { color: #111827; font-size: 13px; }
            QPushButton { 
                background-color: #0056b3; 
                color: white; 
                border: none; 
                padding: 8px 20px; 
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #003d82; }
        """)
        msg.exec()
    
    def handle_view_reports(self):
        """Show report preview dialog with report type selection."""
        # Create dialog
        dialog = QDialog(self.view)
        dialog.setWindowTitle("VIEW REPORTS")
        dialog.setMinimumSize(900, 600)
        dialog.setStyleSheet("QDialog { background-color: white; }")
        
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)
        
        # Title
        title = QLabel("VIEW REPORTS")
        title.setFont(QFont("Arial", 18, QFont.Weight.Bold))
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet("color: #111827; padding-bottom: 10px; border-bottom: 2px solid #d1d5db;")
        layout.addWidget(title)
        
        # Report type selector
        selector_layout = QHBoxLayout()
        selector_label = QLabel("SELECT REPORT TYPE:")
        selector_label.setStyleSheet("color: #111827; font-weight: bold; font-size: 13px;")
        selector_layout.addWidget(selector_label)
        
        report_type_cb = QComboBox()
        report_type_cb.addItems([
            "Stock Summary",
            "Usage Data",
            "Purchasing Trends",
            "Low Stock Alert",
            "Damage Reports",
            "Supplier Performance"
        ])
        report_type_cb.setFixedHeight(35)
        report_type_cb.setStyleSheet("""
            QComboBox {
                border: 2px solid #d1d5db;
                border-radius: 4px;
                padding: 5px 10px;
                background-color: white;
                color: #111827;
                font-size: 13px;
            }
            QComboBox:hover { border-color: #0056b3; }
            QComboBox QAbstractItemView {
                background-color: white;
                color: #111827;
                selection-background-color: #f9fafb;
                selection-color: #111827;
            }
        """)
        selector_layout.addWidget(report_type_cb, 1)
        layout.addLayout(selector_layout)
        
        # Preview table
        preview_table = QTableWidget()
        preview_table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border: 1px solid #d1d5db;
                border-radius: 4px;
                color: #111827;
            }
            QTableWidget::item {
                padding: 8px;
                color: #111827;
            }
            QHeaderView::section {
                background-color: #f9fafb;
                color: #111827;
                font-weight: bold;
                padding: 10px;
                border: none;
                border-bottom: 2px solid #d1d5db;
            }
        """)
        layout.addWidget(preview_table)
        
        # Function to load report preview
        def load_report_preview():
            report_type = report_type_cb.currentText()
            
            try:
                if report_type == "Stock Summary":
                    self.load_stock_summary_preview(preview_table)
                elif report_type == "Usage Data":
                    self.load_usage_data_preview(preview_table)
                elif report_type == "Purchasing Trends":
                    self.load_purchasing_trends_preview(preview_table)
                elif report_type == "Low Stock Alert":
                    self.load_low_stock_preview(preview_table)
                elif report_type == "Damage Reports":
                    self.load_damage_reports_preview(preview_table)
                elif report_type == "Supplier Performance":
                    self.load_supplier_performance_preview(preview_table)
            except Exception as e:
                print(f"Error loading report preview: {e}")
                import traceback
                traceback.print_exc()
        
        # Connect combo box change
        report_type_cb.currentTextChanged.connect(lambda: load_report_preview())
        
        # Close button
        close_btn = QPushButton("CLOSE")
        close_btn.setFixedHeight(40)
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #0056b3;
                color: white;
                border: none;
                border-radius: 4px;
                font-weight: bold;
                font-size: 13px;
            }
            QPushButton:hover {
                background-color: #004494;
            }
        """)
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn)
        
        # Load initial preview
        load_report_preview()
        
        dialog.exec()
    
    def load_stock_summary_preview(self, table):
        """Load stock summary report preview."""
        try:
            item_model = ItemModel()
            items = item_model.list_items()
            
            # Filter by department if Department role
            if self.view.current_role == "Department" and self.view.current_department:
                items = [item for item in items 
                        if item.get('category', '').lower() == self.view.current_department.lower()]
            
            # Group by category
            category_stats = {}
            for item in items:
                category = item.get('category') or "Uncategorized"
                if category not in category_stats:
                    category_stats[category] = {'count': 0, 'value': 0, 'qty': 0}
                category_stats[category]['count'] += 1
                category_stats[category]['value'] += (item.get('stock_qty', 0) * item.get('unit_cost', 0))
                category_stats[category]['qty'] += item.get('stock_qty', 0)
            
            # Set up table
            headers = ['CATEGORY', 'ITEMS COUNT', 'TOTAL QUANTITY', 'TOTAL VALUE']
            table.setColumnCount(len(headers))
            table.setHorizontalHeaderLabels(headers)
            table.setRowCount(len(category_stats))
            
            # Populate table
            for row_idx, (cat, stats) in enumerate(category_stats.items()):
                table.setItem(row_idx, 0, QTableWidgetItem(cat))
                table.setItem(row_idx, 1, QTableWidgetItem(str(stats['count'])))
                table.setItem(row_idx, 2, QTableWidgetItem(str(stats['qty'])))
                table.setItem(row_idx, 3, QTableWidgetItem(f"₱{stats['value']:,.2f}"))
                
                for col in range(4):
                    item = table.item(row_idx, col)
                    if item:
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            
            table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
            
        except Exception as e:
            print(f"Error loading stock summary preview: {e}")
            import traceback
            traceback.print_exc()
    
    def load_usage_data_preview(self, table):
        """Load usage data report preview."""
        try:
            # Get purchases from the last 30 days
            from datetime import datetime, timedelta
            purchases = PurchaseModel.list_purchases()
            
            # Filter recent purchases
            thirty_days_ago = datetime.now() - timedelta(days=30)
            recent_purchases = []
            for p in purchases:
                created_at = p.get('created_at')
                if created_at and isinstance(created_at, str):
                    try:
                        purchase_date = datetime.strptime(created_at[:10], '%Y-%m-%d')
                        if purchase_date >= thirty_days_ago:
                            recent_purchases.append(p)
                    except:
                        pass
                elif isinstance(created_at, datetime):
                    if created_at >= thirty_days_ago:
                        recent_purchases.append(p)
            
            # Set up table
            headers = ['ORDER ID', 'DATE', 'SUPPLIER', 'ITEMS', 'TOTAL AMOUNT', 'STATUS']
            table.setColumnCount(len(headers))
            table.setHorizontalHeaderLabels(headers)
            table.setRowCount(min(len(recent_purchases), 20))  # Show max 20
            
            # Populate table
            for row_idx, purchase in enumerate(recent_purchases[:20]):
                order_id = f"#{purchase.get('id', 0):04d}"
                date_str = str(purchase.get('created_at', ''))[:10] if purchase.get('created_at') else 'N/A'
                supplier = purchase.get('supplier_name', 'N/A')
                items_count = str(purchase.get('items_count', 0)) + ' items'
                total = f"₱{float(purchase.get('total_amount', 0)):,.2f}"
                status = str(purchase.get('status', 'PENDING')).upper()
                
                table.setItem(row_idx, 0, QTableWidgetItem(order_id))
                table.setItem(row_idx, 1, QTableWidgetItem(date_str))
                table.setItem(row_idx, 2, QTableWidgetItem(supplier))
                table.setItem(row_idx, 3, QTableWidgetItem(items_count))
                table.setItem(row_idx, 4, QTableWidgetItem(total))
                table.setItem(row_idx, 5, QTableWidgetItem(status))
                
                for col in range(6):
                    item = table.item(row_idx, col)
                    if item:
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            
            table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
            
        except Exception as e:
            print(f"Error loading usage data preview: {e}")
            import traceback
            traceback.print_exc()
    
    def load_purchasing_trends_preview(self, table):
        """Load purchasing trends report preview."""
        try:
            purchases = PurchaseModel.list_purchases()
            
            # Group by month
            from datetime import datetime
            from collections import defaultdict
            
            monthly_stats = defaultdict(lambda: {'count': 0, 'total': 0})
            
            for p in purchases:
                created_at = p.get('created_at')
                if created_at:
                    if isinstance(created_at, str):
                        try:
                            month_key = created_at[:7]  # YYYY-MM
                        except:
                            continue
                    elif isinstance(created_at, datetime):
                        month_key = created_at.strftime('%Y-%m')
                    else:
                        continue
                    
                    monthly_stats[month_key]['count'] += 1
                    monthly_stats[month_key]['total'] += float(p.get('total_amount', 0))
            
            # Sort by month
            sorted_months = sorted(monthly_stats.items(), reverse=True)[:12]  # Last 12 months
            
            # Set up table
            headers = ['MONTH', 'ORDERS COUNT', 'TOTAL AMOUNT', 'AVG ORDER VALUE']
            table.setColumnCount(len(headers))
            table.setHorizontalHeaderLabels(headers)
            table.setRowCount(len(sorted_months))
            
            # Populate table
            for row_idx, (month, stats) in enumerate(sorted_months):
                avg_value = stats['total'] / stats['count'] if stats['count'] > 0 else 0
                
                table.setItem(row_idx, 0, QTableWidgetItem(month))
                table.setItem(row_idx, 1, QTableWidgetItem(str(stats['count'])))
                table.setItem(row_idx, 2, QTableWidgetItem(f"₱{stats['total']:,.2f}"))
                table.setItem(row_idx, 3, QTableWidgetItem(f"₱{avg_value:,.2f}"))
                
                for col in range(4):
                    item = table.item(row_idx, col)
                    if item:
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            
            table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
            
        except Exception as e:
            print(f"Error loading purchasing trends preview: {e}")
            import traceback
            traceback.print_exc()
    
    def load_low_stock_preview(self, table):
        """Load low stock alert report preview."""
        try:
            item_model = ItemModel()
            items = item_model.list_items()
            
            # Filter by department if Department role
            if self.view.current_role == "Department" and self.view.current_department:
                items = [item for item in items 
                        if item.get('category', '').lower() == self.view.current_department.lower()]
            
            # Find low stock items
            low_stock_items = [item for item in items 
                             if item.get('stock_qty', 0) <= item.get('min_stock', 0)]
            
            # Set up table
            headers = ['ITEM NAME', 'CATEGORY', 'CURRENT STOCK', 'MIN STOCK', 'SHORTAGE', 'STATUS']
            table.setColumnCount(len(headers))
            table.setHorizontalHeaderLabels(headers)
            table.setRowCount(len(low_stock_items))
            
            # Populate table
            for row_idx, item in enumerate(low_stock_items):
                name = item.get('name', 'Unknown')
                category = item.get('category', 'General')
                current_stock = item.get('stock_qty', 0)
                min_stock = item.get('min_stock', 0)
                shortage = max(0, min_stock - current_stock)
                
                if current_stock == 0:
                    status = "OUT OF STOCK"
                elif current_stock <= min_stock * 0.5:
                    status = "CRITICAL"
                else:
                    status = "LOW"
                
                table.setItem(row_idx, 0, QTableWidgetItem(name))
                table.setItem(row_idx, 1, QTableWidgetItem(category))
                table.setItem(row_idx, 2, QTableWidgetItem(str(current_stock)))
                table.setItem(row_idx, 3, QTableWidgetItem(str(min_stock)))
                table.setItem(row_idx, 4, QTableWidgetItem(str(shortage)))
                table.setItem(row_idx, 5, QTableWidgetItem(status))
                
                for col in range(6):
                    item_widget = table.item(row_idx, col)
                    if item_widget:
                        item_widget.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            
            table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
            
        except Exception as e:
            print(f"Error loading low stock preview: {e}")
            import traceback
            traceback.print_exc()
    
    def load_damage_reports_preview(self, table):
        """Load damage reports preview."""
        try:
            damages = DamageModel.list_damages()
            
            # Set up table
            headers = ['DATE', 'ITEM', 'QUANTITY', 'CATEGORY', 'REASON', 'REPORTED BY']
            table.setColumnCount(len(headers))
            table.setHorizontalHeaderLabels(headers)
            table.setRowCount(min(len(damages), 20))  # Show max 20
            
            # Populate table
            for row_idx, damage in enumerate(damages[:20]):
                date_str = str(damage.get('created_at', ''))[:10] if damage.get('created_at') else 'N/A'
                
                # Get item name (would need to join with items table)
                item_name = f"Item #{damage.get('item_id', 'N/A')}"
                quantity = str(damage.get('quantity', 0))
                category = damage.get('category', 'N/A')
                reason = damage.get('reason', 'N/A')
                reported_by = damage.get('created_by', 'N/A')
                
                table.setItem(row_idx, 0, QTableWidgetItem(date_str))
                table.setItem(row_idx, 1, QTableWidgetItem(item_name))
                table.setItem(row_idx, 2, QTableWidgetItem(quantity))
                table.setItem(row_idx, 3, QTableWidgetItem(category))
                table.setItem(row_idx, 4, QTableWidgetItem(reason))
                table.setItem(row_idx, 5, QTableWidgetItem(reported_by))
                
                for col in range(6):
                    item = table.item(row_idx, col)
                    if item:
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            
            table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
            
        except Exception as e:
            print(f"Error loading damage reports preview: {e}")
            import traceback
            traceback.print_exc()
    
    def load_supplier_performance_preview(self, table):
        """Load supplier performance report preview."""
        try:
            from models.purchase import SupplierModel
            suppliers = SupplierModel.list_suppliers()
            purchases = PurchaseModel.list_purchases()
            
            # Calculate stats per supplier
            supplier_stats = {}
            for supplier in suppliers:
                supplier_id = supplier.get('id')
                supplier_stats[supplier_id] = {
                    'name': supplier.get('name', 'Unknown'),
                    'orders': 0,
                    'total_amount': 0,
                    'completed': 0,
                    'pending': 0
                }
            
            for purchase in purchases:
                supplier_id = purchase.get('supplier_id')
                if supplier_id in supplier_stats:
                    supplier_stats[supplier_id]['orders'] += 1
                    supplier_stats[supplier_id]['total_amount'] += float(purchase.get('total_amount', 0))
                    
                    status = purchase.get('status', '').lower()
                    if status in ['completed', 'received']:
                        supplier_stats[supplier_id]['completed'] += 1
                    elif status == 'pending':
                        supplier_stats[supplier_id]['pending'] += 1
            
            # Set up table
            headers = ['SUPPLIER', 'TOTAL ORDERS', 'TOTAL AMOUNT', 'COMPLETED', 'PENDING', 'COMPLETION RATE']
            table.setColumnCount(len(headers))
            table.setHorizontalHeaderLabels(headers)
            table.setRowCount(len(supplier_stats))
            
            # Populate table
            for row_idx, (supplier_id, stats) in enumerate(supplier_stats.items()):
                completion_rate = (stats['completed'] / stats['orders'] * 100) if stats['orders'] > 0 else 0
                
                table.setItem(row_idx, 0, QTableWidgetItem(stats['name']))
                table.setItem(row_idx, 1, QTableWidgetItem(str(stats['orders'])))
                table.setItem(row_idx, 2, QTableWidgetItem(f"₱{stats['total_amount']:,.2f}"))
                table.setItem(row_idx, 3, QTableWidgetItem(str(stats['completed'])))
                table.setItem(row_idx, 4, QTableWidgetItem(str(stats['pending'])))
                table.setItem(row_idx, 5, QTableWidgetItem(f"{completion_rate:.1f}%"))
                
                for col in range(6):
                    item = table.item(row_idx, col)
                    if item:
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            
            table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
            
        except Exception as e:
            print(f"Error loading supplier performance preview: {e}")
            import traceback
            traceback.print_exc()
    
    def handle_usage_history(self):
        """Show usage history report - kept for backwards compatibility."""
        self.handle_view_reports()
    
    def handle_low_stock_alert(self):
        """Show low stock alert report."""
        try:
            item_model = ItemModel()
            items = item_model.list_items()
            
            # Filter by department if Department role
            if self.view.current_role == "Department" and self.view.current_department:
                items = [item for item in items 
                        if item.get('category', '').lower() == self.view.current_department.lower()]
            
            # Find low stock items
            low_stock_items = [item for item in items 
                             if item.get('stock_qty', 0) <= item.get('min_stock', 0)]
            
            if not low_stock_items:
                msg = QMessageBox(self.view)
                msg.setIcon(QMessageBox.Icon.Information)
                msg.setWindowTitle("Low Stock Alert")
                msg.setText("No low stock items found.")
                msg.setStyleSheet("""
                    QMessageBox { background-color: white; }
                    QLabel { color: #111827; font-size: 13px; }
                    QPushButton { 
                        background-color: #10b981; 
                        color: white; 
                        border: none; 
                        padding: 8px 20px; 
                        border-radius: 4px;
                        font-weight: bold;
                    }
                    QPushButton:hover { background-color: #059669; }
                """)
                msg.exec()
            else:
                alert_text = f"Found {len(low_stock_items)} low stock items:\n\n"
                for item in low_stock_items[:5]:  # Show first 5
                    alert_text += f"• {item.get('name')}: {item.get('stock_qty')} {item.get('unit')} (Min: {item.get('min_stock')})\n"
                
                if len(low_stock_items) > 5:
                    alert_text += f"\n... and {len(low_stock_items) - 5} more items"
                
                msg = QMessageBox(self.view)
                msg.setIcon(QMessageBox.Icon.Warning)
                msg.setWindowTitle("Low Stock Alert")
                msg.setText(alert_text)
                msg.setStyleSheet("""
                    QMessageBox { background-color: white; }
                    QLabel { color: #111827; font-size: 13px; }
                    QPushButton { 
                        background-color: #f59e0b; 
                        color: white; 
                        border: none; 
                        padding: 8px 20px; 
                        border-radius: 4px;
                        font-weight: bold;
                    }
                    QPushButton:hover { background-color: #d97706; }
                """)
                msg.exec()
                
        except Exception as e:
            msg = QMessageBox(self.view)
            msg.setIcon(QMessageBox.Icon.Critical)
            msg.setWindowTitle("Error")
            msg.setText(f"Failed to check low stock items:\n{e}")
            msg.setStyleSheet("""
                QMessageBox { background-color: white; }
                QLabel { color: #111827; font-size: 13px; }
                QPushButton { 
                    background-color: #ef4444; 
                    color: white; 
                    border: none; 
                    padding: 8px 20px; 
                    border-radius: 4px;
                    font-weight: bold;
                }
                QPushButton:hover { background-color: #dc2626; }
            """)
            msg.exec()
    
    def handle_export_report(self):
        """Export current report data."""
        msg = QMessageBox(self.view)
        msg.setIcon(QMessageBox.Icon.Information)
        msg.setWindowTitle("Export Report")
        msg.setText("Export functionality coming soon!")
        msg.setStyleSheet("""
            QMessageBox { background-color: white; }
            QLabel { color: #111827; font-size: 13px; }
            QPushButton { 
                background-color: #0056b3; 
                color: white; 
                border: none; 
                padding: 8px 20px; 
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #003d82; }
        """)
        msg.exec()
    
    # Old methods below (kept for compatibility with other roles)
    
    def handle_stock_summary(self):
        """Generate and display stock summary report."""
        try:
            item_model = ItemModel()
            items = item_model.list_items()
            
            # Calculate summary statistics
            total_items = len(items)
            total_value = sum((item.get('stock_qty', 0) * item.get('unit_cost', 0)) for item in items)
            low_stock_count = sum(1 for item in items if item.get('stock_qty', 0) <= item.get('min_stock', 0))
            out_of_stock = sum(1 for item in items if item.get('stock_qty', 0) == 0)
            
            # Group by category
            category_stats = {}
            for item in items:
                category = item.get('category') or "Uncategorized"
                if category not in category_stats:
                    category_stats[category] = {'count': 0, 'value': 0, 'qty': 0}
                category_stats[category]['count'] += 1
                category_stats[category]['value'] += (item.get('stock_qty', 0) * item.get('unit_cost', 0))
                category_stats[category]['qty'] += item.get('stock_qty', 0)
            
            # Prepare report data
            report_data = {
                'title': 'STOCK SUMMARY REPORT',
                'date_range': 'Current Stock Levels',
                'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M'),
                'summary': {
                    'Total Items': total_items,
                    'Total Stock Value': f'₱{total_value:,.2f}',
                    'Low Stock Items': low_stock_count,
                    'Out of Stock Items': out_of_stock,
                    'Categories': len(category_stats)
                },
                'headers': ['CATEGORY', 'ITEMS COUNT', 'TOTAL QUANTITY', 'TOTAL VALUE'],
                'table_data': [
                    [cat, stats['count'], stats['qty'], f"₱{stats['value']:,.2f}"]
                    for cat, stats in category_stats.items()
                ]
            }
            
            # Show report dialog
            dlg = ViewReportDialog(report_data, self.view)
            dlg.exec()
            
        except Exception as e:
            msg = QMessageBox(self.view)
            msg.setIcon(QMessageBox.Icon.Critical)
            msg.setWindowTitle("Error")
            msg.setText(f"Failed to generate stock summary:\n{e}")
            msg.setStyleSheet("QLabel { color: #000000; }")
            msg.exec()
    
    def handle_usage_data(self):
        """Generate and display usage data report."""
        try:
            damage_model = DamageModel()
            damages = damage_model.list_damages()
            
            item_model = ItemModel()
            
            # Calculate usage statistics
            total_damages = len(damages)
            
            # Group by category
            category_damages = {}
            total_qty_damaged = 0
            
            for damage in damages:
                item_id = damage.get('item_id')
                category = damage.get('category', 'Unknown')
                quantity = damage.get('quantity', 0)
                total_qty_damaged += quantity
                
                if category not in category_damages:
                    category_damages[category] = {'count': 0, 'quantity': 0}
                category_damages[category]['count'] += 1
                category_damages[category]['quantity'] += quantity
            
            # Prepare report data
            report_data = {
                'title': 'USAGE DATA REPORT',
                'date_range': 'All Damage Reports',
                'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M'),
                'summary': {
                    'Total Damage Reports': total_damages,
                    'Total Quantity Damaged': total_qty_damaged,
                    'Damage Categories': len(category_damages)
                },
                'headers': ['DAMAGE CATEGORY', 'REPORT COUNT', 'TOTAL QUANTITY'],
                'table_data': [
                    [cat, stats['count'], stats['quantity']]
                    for cat, stats in category_damages.items()
                ]
            }
            
            # Show report dialog
            dlg = ViewReportDialog(report_data, self.view)
            dlg.exec()
            
        except Exception as e:
            msg = QMessageBox(self.view)
            msg.setIcon(QMessageBox.Icon.Critical)
            msg.setWindowTitle("Error")
            msg.setText(f"Failed to generate usage data:\n{e}")
            msg.setStyleSheet("QLabel { color: #000000; }")
            msg.exec()
    
    def handle_purchasing_trends(self):
        """Generate and display purchasing trends report."""
        try:
            purchases = self.model.list_purchases()
            
            # Calculate trends statistics
            total_purchases = len(purchases)
            total_amount = sum(p.get('total_amount', 0) for p in purchases)
            
            # Group by status
            status_stats = {}
            for purchase in purchases:
                status = purchase.get('status', 'Unknown').capitalize()
                amount = purchase.get('total_amount', 0)
                
                if status not in status_stats:
                    status_stats[status] = {'count': 0, 'amount': 0}
                status_stats[status]['count'] += 1
                status_stats[status]['amount'] += amount
            
            # Group by month
            monthly_stats = {}
            for purchase in purchases:
                created_at = purchase.get('created_at')
                if created_at:
                    if isinstance(created_at, str):
                        try:
                            purchase_date = datetime.strptime(created_at, "%Y-%m-%d %H:%M:%S")
                        except ValueError:
                            try:
                                purchase_date = datetime.strptime(created_at, "%Y-%m-%d")
                            except ValueError:
                                continue
                    else:
                        purchase_date = created_at
                    
                    month_key = purchase_date.strftime('%Y-%m')
                    amount = purchase.get('total_amount', 0)
                    
                    if month_key not in monthly_stats:
                        monthly_stats[month_key] = {'count': 0, 'amount': 0}
                    monthly_stats[month_key]['count'] += 1
                    monthly_stats[month_key]['amount'] += amount
            
            # Prepare report data
            report_data = {
                'title': 'PURCHASING TRENDS REPORT',
                'date_range': 'All Purchase Orders',
                'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M'),
                'summary': {
                    'Total Purchase Orders': total_purchases,
                    'Total Amount': f'₱{total_amount:,.2f}',
                    'Average Order Value': f'₱{(total_amount/total_purchases if total_purchases > 0 else 0):,.2f}'
                },
                'headers': ['MONTH', 'ORDER COUNT', 'TOTAL AMOUNT'],
                'table_data': [
                    [month, stats['count'], f"₱{stats['amount']:,.2f}"]
                    for month, stats in sorted(monthly_stats.items(), reverse=True)
                ]
            }
            
            # Show report dialog
            dlg = ViewReportDialog(report_data, self.view)
            dlg.exec()
            
        except Exception as e:
            msg = QMessageBox(self.view)
            msg.setIcon(QMessageBox.Icon.Critical)
            msg.setWindowTitle("Error")
            msg.setText(f"Failed to generate purchasing trends:\n{e}")
            msg.setStyleSheet("QLabel { color: #000000; }")
            msg.exec()
    
    def handle_export_report(self):
        """Export report to CSV, Excel, or PDF file."""
        print("Export button clicked!")  # Debug
        try:
            # First, ask user to select report type and format
            dlg = GenerateReportDialog(self.view)
            print("Dialog created")  # Debug
            
            # Wait for user to click generate or cancel
            if dlg.exec():
                print("Dialog accepted")  # Debug
                data = dlg.get_data()
                report_type = data['report_type']
                format_type = data['format']
                print(f"Report type: {report_type}, Format: {format_type}")  # Debug
                
                # Set file extension based on format
                if format_type == "CSV":
                    file_filter = "CSV Files (*.csv)"
                    default_ext = ".csv"
                elif format_type == "Excel":
                    file_filter = "Excel Files (*.xlsx)"
                    default_ext = ".xlsx"
                elif format_type == "PDF":
                    file_filter = "PDF Files (*.pdf)"
                    default_ext = ".pdf"
                else:
                    file_filter = "All Files (*.*)"
                    default_ext = ""
                
                # Ask user where to save the file
                default_filename = f"{report_type.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{default_ext}"
                file_path, _ = QFileDialog.getSaveFileName(
                    self.view,
                    "Export Report",
                    default_filename,
                    file_filter
                )
                print(f"File path: {file_path}")  # Debug
                
                # If user selected a path, generate and save the report
                if file_path:
                    # Generate the appropriate report data
                    report_data = None
                    if report_type == "Stock Summary":
                        report_data = self.generate_stock_summary_data()
                    elif report_type == "Usage Data":
                        report_data = self.generate_usage_data()
                    elif report_type == "Purchasing Trends":
                        report_data = self.generate_purchasing_trends_data()
                    elif report_type == "Low Stock Alert":
                        report_data = self.generate_low_stock_data()
                    elif report_type == "Damage Reports":
                        report_data = self.generate_usage_data()  # Same as usage data
                    elif report_type == "Supplier Performance":
                        report_data = self.generate_supplier_performance_data()
                    
                    if report_data:
                        # Save based on format
                        if format_type == "CSV":
                            self.export_to_csv(file_path, report_type, report_data)
                        elif format_type == "Excel":
                            self.export_to_excel(file_path, report_type, report_data)
                        elif format_type == "PDF":
                            self.export_to_pdf(file_path, report_type, report_data)
                        
                        msg = QMessageBox(self.view)
                        msg.setIcon(QMessageBox.Icon.Information)
                        msg.setWindowTitle("Success")
                        msg.setText(f"Report exported successfully to {format_type}!\n\nSaved to:\n{file_path}")
                        msg.setStyleSheet("QLabel { color: #000000; }")
                        msg.exec()
                        
        except Exception as e:
            msg = QMessageBox(self.view)
            msg.setIcon(QMessageBox.Icon.Critical)
            msg.setWindowTitle("Error")
            msg.setText(f"Failed to export report:\n{e}")
            msg.setStyleSheet("QLabel { color: #000000; }")
            import traceback
            traceback.print_exc()
            msg.exec()
    
    def generate_stock_summary_data(self):
        """Generate stock summary data for export."""
        item_model = ItemModel()
        items = item_model.list_items()
        
        return {
            'headers': ['Item Name', 'SKU', 'Category', 'Stock Qty', 'Unit Cost', 'Total Value', 'Min Stock', 'Status'],
            'table_data': [
                [
                    item.get('name', ''),
                    item.get('sku', ''),
                    item.get('category', 'Uncategorized'),
                    item.get('stock_qty', 0),
                    item.get('unit_cost', 0),
                    item.get('stock_qty', 0) * item.get('unit_cost', 0),
                    item.get('min_stock', 0),
                    'Low Stock' if item.get('stock_qty', 0) <= item.get('min_stock', 0) else 'In Stock'
                ]
                for item in items
            ]
        }
    
    def generate_usage_data(self):
        """Generate usage data for export."""
        damage_model = DamageModel()
        damages = damage_model.list_damages()
        
        return {
            'headers': ['Damage ID', 'Purchase ID', 'Category', 'Quantity', 'Reason', 'Status', 'Created By', 'Date'],
            'table_data': [
                [
                    damage.get('id', ''),
                    damage.get('purchase_id', ''),
                    damage.get('category', ''),
                    damage.get('quantity', 0),
                    damage.get('reason', ''),
                    damage.get('status', ''),
                    damage.get('created_by', ''),
                    damage.get('created_at', '')
                ]
                for damage in damages
            ]
        }
    
    def generate_purchasing_trends_data(self):
        """Generate purchasing trends data for export."""
        purchases = self.model.list_purchases()
        
        return {
            'headers': ['Order ID', 'Supplier', 'Status', 'Total Amount', 'Expected Date', 'Created By', 'Created At'],
            'table_data': [
                [
                    purchase.get('id', ''),
                    purchase.get('supplier_name', ''),
                    purchase.get('status', ''),
                    purchase.get('total_amount', 0),
                    purchase.get('expected_date', ''),
                    purchase.get('created_by', ''),
                    purchase.get('created_at', '')
                ]
                for purchase in purchases
            ]
        }
    
    def generate_low_stock_data(self):
        """Generate low stock alert data for export."""
        item_model = ItemModel()
        items = item_model.list_items()
        
        # Filter only low stock items
        low_stock_items = [item for item in items if item.get('stock_qty', 0) <= item.get('min_stock', 0)]
        
        return {
            'headers': ['Item Name', 'SKU', 'Category', 'Current Stock', 'Min Stock', 'Reorder Needed', 'Unit Cost'],
            'table_data': [
                [
                    item.get('name', ''),
                    item.get('sku', ''),
                    item.get('category', 'Uncategorized'),
                    item.get('stock_qty', 0),
                    item.get('min_stock', 0),
                    max(0, item.get('min_stock', 0) - item.get('stock_qty', 0)),
                    f"₱{item.get('unit_cost', 0):,.2f}"
                ]
                for item in low_stock_items
            ]
        }
    
    def generate_supplier_performance_data(self):
        """Generate supplier performance data for export."""
        from models.purchase import SupplierModel
        supplier_model = SupplierModel()
        suppliers = supplier_model.list_suppliers()
        purchases = self.model.list_purchases()
        
        # Calculate stats per supplier
        supplier_stats = {}
        for purchase in purchases:
            supplier_id = purchase.get('supplier_id')
            if supplier_id:
                if supplier_id not in supplier_stats:
                    supplier_stats[supplier_id] = {
                        'total_orders': 0,
                        'total_amount': 0,
                        'delivered': 0,
                        'pending': 0,
                        'cancelled': 0
                    }
                supplier_stats[supplier_id]['total_orders'] += 1
                supplier_stats[supplier_id]['total_amount'] += purchase.get('total_amount', 0)
                
                status = purchase.get('status', '').lower()
                if status == 'delivered':
                    supplier_stats[supplier_id]['delivered'] += 1
                elif status == 'pending':
                    supplier_stats[supplier_id]['pending'] += 1
                elif status == 'cancelled':
                    supplier_stats[supplier_id]['cancelled'] += 1
        
        return {
            'headers': ['Supplier Name', 'Contact Person', 'Total Orders', 'Total Amount', 'Delivered', 'Pending', 'Cancelled'],
            'table_data': [
                [
                    supplier.get('name', ''),
                    supplier.get('contact_name', ''),
                    supplier_stats.get(supplier.get('id'), {}).get('total_orders', 0),
                    f"₱{supplier_stats.get(supplier.get('id'), {}).get('total_amount', 0):,.2f}",
                    supplier_stats.get(supplier.get('id'), {}).get('delivered', 0),
                    supplier_stats.get(supplier.get('id'), {}).get('pending', 0),
                    supplier_stats.get(supplier.get('id'), {}).get('cancelled', 0)
                ]
                for supplier in suppliers
            ]
        }
    
    def export_to_csv(self, file_path, report_type, report_data):
        """Export report data to CSV file."""
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            # Write title and metadata
            writer.writerow([f"Report Type: {report_type}"])
            writer.writerow([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
            writer.writerow([])  # Empty row
            # Write headers and data
            writer.writerow(report_data['headers'])
            writer.writerows(report_data['table_data'])
    
    def export_to_excel(self, file_path, report_type, report_data):
        """Export report data to Excel file."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        
        # Title
        ws['A1'] = f"Report Type: {report_type}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=10, italic=True)
        
        # Headers (starting at row 4)
        header_fill = PatternFill(start_color="0056b3", end_color="0056b3", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, header in enumerate(report_data['headers'], 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for row_idx, row_data in enumerate(report_data['table_data'], 5):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(file_path)
    
    def export_to_pdf(self, file_path, report_type, report_data):
        """Export report data to PDF file."""
        doc = SimpleDocTemplate(file_path, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title = Paragraph(f"<b>{report_type}</b>", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.2 * inch))
        
        # Metadata
        meta = Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal'])
        elements.append(meta)
        elements.append(Spacer(1, 0.3 * inch))
        
        # Prepare table data
        table_data = [report_data['headers']] + report_data['table_data']
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            # Header row style
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0056b3')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            # Data rows style
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ]))
        
        elements.append(table)
        doc.build(elements)

